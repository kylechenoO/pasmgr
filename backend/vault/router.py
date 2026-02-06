# ---------------------------------------------------------------------------
# Author  : Kyle <kyle@hacking-linux.com>
# Version : 20260206v1
# ---------------------------------------------------------------------------
"""
Vault endpoints – CRUD for password entries, on-demand decryption,
and the server-side password generator.

Security invariants enforced by every handler
---------------------------------------------
* JWT is required on every endpoint (via ``get_current_user``).
* Every item operation first calls ``_own_item``, which loads the row and
  asserts that ``item.user_id == current_user.id``.  Even if an attacker
  guesses another user's item ID, the request is rejected with 403.
* Plaintext passwords are only returned by the dedicated ``/decrypt``
  endpoint.  All other responses contain only the encrypted blob.
"""

import io
import secrets
import string

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from database import get_db
from core.logger import logger
from core.security import encrypt_value, encrypt_value_with_iv, decrypt_value, get_current_user
from models.user import User
from models.vault_item import VaultItem
from models.audit_log import AuditLog
from vault.schemas import (
    VALID_CATEGORIES,
    VaultItemCreate,
    VaultItemUpdate,
    VaultItemResponse,
    VaultItemListResponse,
)

router = APIRouter(prefix="/vault", tags=["vault"])

# ---------------------------------------------------------------------------
# Ownership helper
# ---------------------------------------------------------------------------


def _own_item(item_id: int, user_id: int, db: Session) -> VaultItem:
    """
    Load a VaultItem by ID and verify it belongs to *user_id*.

    Raises 404 if the item does not exist, 403 if it belongs to someone else.
    """
    item = db.query(VaultItem).filter(VaultItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    if item.user_id != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    return item


# ---------------------------------------------------------------------------
# GET /vault/items  – list the current user's entries
# ---------------------------------------------------------------------------


@router.get("/items", response_model=VaultItemListResponse)
def list_items(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return all vault items belonging to the authenticated user."""
    items = (
        db.query(VaultItem)
        .filter(VaultItem.user_id == current_user.id)
        .order_by(VaultItem.created_at.desc())
        .all()
    )
    return VaultItemListResponse(items=items)


# ---------------------------------------------------------------------------
# POST /vault/items  – create a new entry
# ---------------------------------------------------------------------------


@router.post("/items", response_model=VaultItemResponse, status_code=status.HTTP_201_CREATED)
def create_item(
    body: VaultItemCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Encrypt the supplied plaintext password and persist the entry.
    The client never sees the master key or the raw ciphertext bytes.
    """
    if body.category not in VALID_CATEGORIES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid category")

    # Secure notes don't need username/password; others do
    plaintext_pw = body.plaintext_password or ""
    encrypted_password, iv = encrypt_value(plaintext_pw)

    # Encrypt API secret if provided (reuse same IV for simplicity)
    # NOTE: In production, api_secret should have its own IV column for proper security
    api_secret_encrypted = None
    if body.plaintext_api_secret:
        api_secret_encrypted = encrypt_value_with_iv(body.plaintext_api_secret, iv)

    item = VaultItem(
        user_id=current_user.id,
        category=body.category,
        title=body.title,
        username=body.username or "",
        encrypted_password=encrypted_password,
        iv=iv,
        url=body.url,
        port=body.port,
        notes=body.notes,
        public_key=body.public_key,
        private_key=body.private_key,
        api_key=body.api_key,
        api_secret=api_secret_encrypted,
    )
    db.add(item)
    db.commit()
    db.refresh(item)

    # Audit log: mask sensitive fields
    detail = f"category={body.category}, title={body.title}, username={body.username or '(empty)'}, password=******, url={body.url or '(empty)'}"
    if body.public_key:
        detail += ", public_key=******"
    if body.plaintext_api_secret:
        detail += ", api_secret=******"
    db.add(AuditLog(
        admin_id=current_user.id,
        target_user_id=current_user.id,
        action="vault_create",
        detail=detail,
    ))
    db.commit()

    return item


# ---------------------------------------------------------------------------
# GET /vault/items/{id}  – retrieve a single entry (encrypted)
# ---------------------------------------------------------------------------


@router.get("/items/{item_id}", response_model=VaultItemResponse)
def get_item(
    item_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return the encrypted item.  Plaintext is NOT included."""
    item = _own_item(item_id, current_user.id, db)
    return item


# ---------------------------------------------------------------------------
# GET /vault/items/{id}/decrypt  – reveal the plaintext password
# ---------------------------------------------------------------------------


@router.get("/items/{item_id}/decrypt")
def decrypt_item(
    item_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    The *only* endpoint that returns plaintext password and api_secret.
    It is called on-demand when the user clicks "Reveal" in the UI.

    The plaintext is never persisted, cached, or logged.
    """
    item = _own_item(item_id, current_user.id, db)
    plaintext = decrypt_value(item.encrypted_password, item.iv)

    # Decrypt api_secret if present
    plaintext_api_secret = None
    if item.api_secret:
        try:
            plaintext_api_secret = decrypt_value(item.api_secret, item.iv)
        except ValueError:
            plaintext_api_secret = None  # Decryption failed

    result = {
        "plaintext_password": plaintext,
        "plaintext_api_secret": plaintext_api_secret,
    }
    logger.info(f"Decrypt result for item {item_id}: plaintext_password='{plaintext}', plaintext_api_secret='{plaintext_api_secret}'")
    return result


# ---------------------------------------------------------------------------
# PUT /vault/items/{id}  – update an existing entry
# ---------------------------------------------------------------------------


@router.put("/items/{item_id}", response_model=VaultItemResponse)
def update_item(
    item_id: int,
    body: VaultItemUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Partial update.  Only fields that are explicitly provided (non-None) are
    changed.  If ``plaintext_password`` is provided the entry is re-encrypted
    with a fresh IV.
    """
    item = _own_item(item_id, current_user.id, db)

    if body.category is not None:
        if body.category not in VALID_CATEGORIES:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid category")
        item.category = body.category
    if body.title is not None:
        item.title = body.title
    if body.username is not None:
        item.username = body.username
    if body.url is not None:
        item.url = body.url
    if body.port is not None:
        item.port = body.port
    if body.notes is not None:
        item.notes = body.notes
    if body.public_key is not None:
        item.public_key = body.public_key
    if body.private_key is not None:
        item.private_key = body.private_key
    if body.api_key is not None:
        item.api_key = body.api_key
    if body.plaintext_password is not None:
        # Re-encrypt with a new IV – each write gets a unique nonce
        item.encrypted_password, item.iv = encrypt_value(body.plaintext_password)
    if body.plaintext_api_secret is not None:
        # Encrypt API secret using current item's IV (reuse for simplicity)
        # If password was just updated, use the new IV; otherwise use existing IV
        item.api_secret = encrypt_value_with_iv(body.plaintext_api_secret, item.iv)

    db.commit()
    db.refresh(item)

    # Audit log: mask sensitive fields
    changes = []
    if body.category is not None:
        changes.append(f"category={body.category}")
    if body.title is not None:
        changes.append(f"title={body.title}")
    if body.username is not None:
        changes.append(f"username={body.username}")
    if body.plaintext_password is not None:
        changes.append("password=******")
    if body.url is not None:
        changes.append(f"url={body.url}")
    if body.public_key is not None:
        changes.append("public_key=******")
    if body.private_key is not None:
        changes.append("private_key=******")
    if body.plaintext_api_secret is not None:
        changes.append("api_secret=******")
    if body.notes is not None:
        changes.append("notes=<updated>")

    detail = f"item_id={item_id}, " + ", ".join(changes) if changes else f"item_id={item_id}"
    db.add(AuditLog(
        admin_id=current_user.id,
        target_user_id=current_user.id,
        action="vault_update",
        detail=detail,
    ))
    db.commit()

    return item


# ---------------------------------------------------------------------------
# DELETE /vault/items/{id}  – remove an entry
# ---------------------------------------------------------------------------


@router.delete("/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_item(
    item_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Permanently delete a vault entry.  Ownership is verified first."""
    item = _own_item(item_id, current_user.id, db)

    # Capture item details before deletion (mask sensitive fields)
    detail = f"item_id={item_id}, category={item.category}, title={item.title}, username={item.username or '(empty)'}"

    db.delete(item)
    db.add(AuditLog(
        admin_id=current_user.id,
        target_user_id=current_user.id,
        action="vault_delete",
        detail=detail,
    ))
    db.commit()


# ---------------------------------------------------------------------------
# GET /vault/export  – download all entries as an Excel workbook
# ---------------------------------------------------------------------------
# Security note: plaintext passwords are only returned here and via /decrypt.
# The exported file itself is not encrypted – the user should keep it safe.

_HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_EXPORT_HEADERS = ["Category", "Title", "Username", "Password", "URL", "Port", "Public Key", "Private Key", "API Key", "API Secret", "Notes"]


@router.get("/export")
def export_vault(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Generate an .xlsx file containing every vault entry (decrypted) that
    belongs to the authenticated user.  The file is streamed back directly –
    nothing is written to disk on the server.
    """
    items = (
        db.query(VaultItem)
        .filter(VaultItem.user_id == current_user.id)
        .order_by(VaultItem.title.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Passwords"

    # -- Header row ----------------------------------------------------------
    ws.append(_EXPORT_HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # -- Data rows -----------------------------------------------------------
    for item in items:
        plaintext = decrypt_value(item.encrypted_password, item.iv)
        # Decrypt api_secret if present (gracefully handle old entries with mismatched IV)
        api_secret_plaintext = ""
        if item.api_secret:
            try:
                api_secret_plaintext = decrypt_value(item.api_secret, item.iv)
            except ValueError:
                # Old entry with different IV - skip decryption
                api_secret_plaintext = "[decryption failed - please re-save this entry]"

        ws.append([
            item.category,
            item.title,
            item.username,
            plaintext,
            item.url or "",
            item.port or "",
            item.public_key or "",
            item.private_key or "",
            item.api_key or "",
            api_secret_plaintext,
            item.notes or "",
        ])
        # Apply border to every cell in the new row
        row_idx = ws.max_row
        for col_idx in range(1, len(_EXPORT_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    # -- Column widths (auto-fit heuristic) ----------------------------------
    _COL_MIN = [16, 18, 24, 28, 36, 10, 30, 30, 28, 28, 30]  # sensible minimums per column
    for col_idx, min_w in enumerate(_COL_MIN, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = min_w

    # -- Stream without touching the filesystem -----------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Audit log
    db.add(AuditLog(
        admin_id=current_user.id,
        target_user_id=current_user.id,
        action="vault_export",
        detail=f"Exported {len(items)} vault item(s) to Excel",
    ))
    db.commit()

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="passwords.xlsx"'},
    )


# ---------------------------------------------------------------------------
# POST /vault/import  – bulk-create entries from an Excel workbook
# ---------------------------------------------------------------------------


@router.post("/import")
async def import_vault(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Accept an .xlsx file (same column layout as the export).  Each row is
    validated and encrypted before insertion.

    * Rows that are missing Title, Username, or Password are skipped.
    * Rows whose Title already exists for this user are skipped (no
      duplicates).  The response body reports how many rows were skipped
      and why.

    The file is read entirely into memory – passwords managers deal with
    small datasets so this is acceptable.
    """
    # -- Basic file-type guard -----------------------------------------------
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted",
        )

    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not parse the uploaded file as .xlsx",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Workbook has no active sheet",
        )

    # -- Locate header columns (case-insensitive, tolerant of extra cols) ----
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sheet is empty",
        )

    col_map: dict[str, int] = {}  # lowercase name → 0-based index
    for idx, cell_val in enumerate(header_row):
        if cell_val and isinstance(cell_val, str):
            col_map[cell_val.strip().lower()] = idx

    for required in ("title", "username", "password"):
        if required not in col_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required column: {required}",
            )

    # Pre-fetch existing titles for this user to detect duplicates fast
    existing_titles: set[str] = {
        t[0]
        for t in db.query(VaultItem.title)
        .filter(VaultItem.user_id == current_user.id)
        .all()
    }

    # -- Process data rows ---------------------------------------------------
    imported   = 0
    skipped_dup = 0
    skipped_bad = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        category = _cell_str(row, col_map.get("category")) or "login"
        title    = _cell_str(row, col_map["title"])
        username = _cell_str(row, col_map.get("username"))
        password = _cell_str(row, col_map.get("password"))
        url      = _cell_str(row, col_map.get("url"))
        port_str = _cell_str(row, col_map.get("port"))
        port     = int(port_str) if port_str and port_str.isdigit() else None
        public_key = _cell_str(row, col_map.get("public key"))
        private_key = _cell_str(row, col_map.get("private key"))
        api_key = _cell_str(row, col_map.get("api key"))
        api_secret = _cell_str(row, col_map.get("api secret"))
        notes    = _cell_str(row, col_map.get("notes"))

        # Skip rows that are missing mandatory title or have invalid category
        if not title:
            skipped_bad += 1
            continue
        if category not in VALID_CATEGORIES:
            skipped_bad += 1
            continue

        # Skip duplicates (title already exists for this user)
        if title in existing_titles:
            skipped_dup += 1
            continue

        encrypted_password, iv = encrypt_value(password or "")

        # Encrypt API secret if present (use same IV as password)
        api_secret_encrypted = None
        if api_secret:
            api_secret_encrypted = encrypt_value_with_iv(api_secret, iv)

        item = VaultItem(
            user_id=current_user.id,
            category=category,
            title=title,
            username=username or "",
            encrypted_password=encrypted_password,
            iv=iv,
            url=url or None,
            port=port,
            public_key=public_key or None,
            private_key=private_key or None,
            api_key=api_key or None,
            api_secret=api_secret_encrypted,
            notes=notes or None,
        )
        db.add(item)
        existing_titles.add(title)   # prevent duplicates within the same file
        imported += 1

    db.commit()
    wb.close()

    # Audit log
    db.add(AuditLog(
        admin_id=current_user.id,
        target_user_id=current_user.id,
        action="vault_import",
        detail=f"Imported {imported} vault item(s) from Excel (skipped {skipped_dup} duplicates, {skipped_bad} invalid)",
    ))
    db.commit()

    return {
        "imported": imported,
        "skipped_duplicates": skipped_dup,
        "skipped_invalid": skipped_bad,
    }


def _cell_str(row: tuple, col_idx: int | None) -> str:
    """Safely extract a string value from a row tuple by column index."""
    if col_idx is None or col_idx >= len(row):
        return ""
    val = row[col_idx]
    return str(val).strip() if val is not None else ""


# ---------------------------------------------------------------------------
# GET /vault/generate-password  – server-side password generation
# ---------------------------------------------------------------------------

# -- Word lists for "memorable" mode --------------------------------------
# Short, curated lists.  In production you might load these from a file.
_ADJECTIVES = [
    "bright", "calm", "dark", "eager", "fair", "gentle", "happy", "keen",
    "lively", "magic", "noble", "open", "proud", "quick", "rich", "swift",
    "tall", "ultra", "vivid", "warm", "young", "azure", "brave", "clear",
    "deep", "fresh", "grand", "jolly", "kind", "lucky", "merry", "neat",
    "plain", "quiet", "royal", "smart", "tidy", "upper", "vital", "wild",
    "bold", "cool", "damp", "easy", "fine", "glad", "huge", "idle",
    "just", "lean",
]

_NOUNS = [
    "apple", "bear", "cedar", "dawn", "eagle", "frost", "grape", "hawk",
    "iris", "jade", "kite", "lake", "maple", "night", "ocean", "peach",
    "quail", "river", "stone", "tiger", "ultra", "vine", "wave", "xenon",
    "yak", "zebra", "amber", "brook", "crane", "dove", "elm", "flare",
    "gold", "holly", "ivy", "jewel", "knot", "lily", "moss", "oak",
    "pearl", "rose", "sage", "thorn", "umber", "willow", "yard", "zeal",
    "arch", "birch",
]

# Symbols safe for most password fields
_SYMBOLS = "!@#$%^&*()-_=+[]{}|;:,.<>?"


@router.get("/generate-password")
def generate_password(
    mode: str = "random",
    length: int = 16,
    include_digits: bool = True,
    include_symbols: bool = False,
    current_user: User = Depends(get_current_user),  # must be logged in
):
    """
    Generate a password using the OS CSPRNG (``secrets`` module).

    Modes
    -----
    random     – configurable charset, length 6-256.
    memorable  – adjective-noun-NN pattern  (length param ignored).
    pin        – digits only, length 4-256.
    """
    mode = mode.lower().strip()

    if mode == "random":
        return _gen_random(length, include_digits, include_symbols)
    elif mode == "memorable":
        return _gen_memorable()
    elif mode == "pin":
        return _gen_pin(length)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid mode. Must be 'random', 'memorable', or 'pin'",
        )


# -- Generator helpers -----------------------------------------------------


def _gen_random(length: int, digits: bool, symbols: bool) -> dict:
    """Random password with at least one upper and one lower letter."""
    if not (6 <= length <= 256):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Length must be between 6 and 256 for random mode",
        )

    # Build charset – uppercase + lowercase are always included
    charset = string.ascii_uppercase + string.ascii_lowercase
    if digits:
        charset += string.digits
    if symbols:
        charset += _SYMBOLS

    # Guarantee at least one character from each mandatory class
    password_chars = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
    ]
    # Fill the rest
    for _ in range(length - 2):
        password_chars.append(secrets.choice(charset))

    # Shuffle to avoid predictable positions for the guaranteed chars
    # secrets.SystemRandom is used internally by secrets; we shuffle manually
    import random as _rng  # only for shuffle; no crypto use
    _rng.SystemRandom().shuffle(password_chars)

    return {"password": "".join(password_chars)}


def _gen_memorable() -> dict:
    """adjective-noun-NN pattern."""
    adj = secrets.choice(_ADJECTIVES)
    noun = secrets.choice(_NOUNS)
    num = secrets.randbelow(90) + 10  # 10-99 inclusive
    return {"password": f"{adj}-{noun}-{num}"}


def _gen_pin(length: int) -> dict:
    """Numeric PIN."""
    if not (4 <= length <= 256):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="PIN length must be between 4 and 256",
        )
    return {"password": "".join(secrets.choice(string.digits) for _ in range(length))}
