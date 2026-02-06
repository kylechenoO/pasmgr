# ---------------------------------------------------------------------------
# Author  : Kyle <kyle@hacking-linux.com>
# Version : 20260206v1
# ---------------------------------------------------------------------------
"""
Admin endpoints – user lifecycle management.

Every endpoint in this router is guarded by ``require_admin``.  A request
that carries a valid JWT but belongs to a ``user`` role will receive 403
before any business logic runs.
"""

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, aliased
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_db
from core.security import hash_password, require_admin
from models.user import User
from models.audit_log import AuditLog
from admin.schemas import (
    AuditLogListResponse,
    AuditLogRow,
    ChangeRoleRequest,
    CreateUserRequest,
    ResetPasswordRequest,
    UserRow,
    UserListResponse,
)

router = APIRouter(prefix="/admin", tags=["admin"])

_VALID_ROLES = {"admin", "user"}


# ---------------------------------------------------------------------------
# POST /admin/users  – create a new user
# ---------------------------------------------------------------------------


@router.post("/users", response_model=UserRow, status_code=status.HTTP_201_CREATED)
def create_user(
    body: CreateUserRequest,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Create a user account.  The new user will have ``force_password_change``
    set to True so they must set their own password on first login.
    """
    if body.role not in _VALID_ROLES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid role. Must be 'admin' or 'user'",
        )

    # Uniqueness check
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Email already exists",
        )

    password_hash, salt = hash_password(body.password)
    user = User(
        email=body.email,
        password_hash=password_hash,
        salt=salt,
        role=body.role,
        is_active=True,
        force_password_change=True,  # must change on first login
    )
    db.add(user)
    db.flush()  # get user.id before commit
    db.add(AuditLog(admin_id=admin.id, target_user_id=user.id, action="create_user", detail=f"role={body.role}"))
    db.commit()
    db.refresh(user)
    return user


# ---------------------------------------------------------------------------
# GET /admin/users  – list all users
# ---------------------------------------------------------------------------


@router.get("/users", response_model=UserListResponse)
def list_users(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return every user row (no password data – handled by the schema)."""
    users = db.query(User).order_by(User.id).all()
    return UserListResponse(users=users)


# ---------------------------------------------------------------------------
# PUT /admin/users/{id}/reset-password  – admin resets another user's password
# ---------------------------------------------------------------------------


@router.put("/users/{user_id}/reset-password")
def reset_password(
    user_id: int,
    body: ResetPasswordRequest,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Overwrite a user's password.  ``force_password_change`` is set back to
    True so the user must pick a new password on their next login.
    """
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    new_hash, new_salt = hash_password(body.new_password)
    target.password_hash = new_hash
    target.salt = new_salt
    target.force_password_change = True
    db.add(AuditLog(admin_id=admin.id, target_user_id=user_id, action="reset_password"))
    db.commit()

    return {"detail": "Password reset successfully"}


# ---------------------------------------------------------------------------
# PUT /admin/users/{id}/disable  – soft-disable a user account
# ---------------------------------------------------------------------------


@router.put("/users/{user_id}/disable")
def disable_user(
    user_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Set ``is_active = False``.  The user can no longer log in, and any
    existing tokens will be rejected by ``get_current_user``.

    Guard: an admin cannot disable their own account.
    """
    if user_id == admin.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot disable yourself",
        )

    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    target.is_active = False
    db.add(AuditLog(admin_id=admin.id, target_user_id=user_id, action="disable_user"))
    db.commit()

    return {"detail": "User disabled"}


# ---------------------------------------------------------------------------
# PUT /admin/users/{id}/enable  – re-activate a disabled user account
# ---------------------------------------------------------------------------


@router.put("/users/{user_id}/enable")
def enable_user(
    user_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Set ``is_active = True`` so the user can log in again."""
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    target.is_active = True
    db.add(AuditLog(admin_id=admin.id, target_user_id=user_id, action="enable_user"))
    db.commit()

    return {"detail": "User enabled"}


# ---------------------------------------------------------------------------
# PUT /admin/users/{id}/change-role  – promote or demote a user
# ---------------------------------------------------------------------------


@router.put("/users/{user_id}/change-role")
def change_role(
    user_id: int,
    body: ChangeRoleRequest,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Change the role of an existing user.  Guards:
    * Role value must be 'admin' or 'user'.
    * An admin cannot change their own role (prevents accidental self-lockout).
    """
    if body.role not in _VALID_ROLES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid role. Must be 'admin' or 'user'",
        )

    if user_id == admin.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot change your own role",
        )

    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    target.role = body.role
    db.add(AuditLog(admin_id=admin.id, target_user_id=user_id, action="change_role", detail=f"new_role={body.role}"))
    db.commit()

    return {"detail": "Role updated"}


# ---------------------------------------------------------------------------
# GET /admin/audit-logs  – paginated audit trail with optional filters
# ---------------------------------------------------------------------------


@router.get("/audit-logs", response_model=AuditLogListResponse)
def list_audit_logs(
    emails: list[str] | None = Query(None, description="Filter by exact email(s) – repeated param"),
    since: datetime | None = Query(None, description="ISO-8601 start of time window"),
    until: datetime | None = Query(None, description="ISO-8601 end of time window"),
    limit: int = Query(200, ge=1, le=1000),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Return audit log rows newest-first.  Supports optional filters:

    * ``emails`` – one or more exact email addresses; match rows where
                   *either* admin_id or target_user_id belongs to one of them.
    * ``since`` / ``until`` – ISO-8601 bounds on ``created_at``.
    * ``limit`` – max rows returned (default 200, cap 1000).
    """
    AdminUser  = aliased(User)
    TargetUser = aliased(User)

    q = (
        db.query(AuditLog)
        .outerjoin(AdminUser,  AuditLog.admin_id       == AdminUser.id)
        .outerjoin(TargetUser, AuditLog.target_user_id == TargetUser.id)
    )

    if emails:
        q = q.filter(
            AdminUser.email.in_(emails) | TargetUser.email.in_(emails)
        )
    if since:
        q = q.filter(AuditLog.created_at >= since)
    if until:
        q = q.filter(AuditLog.created_at <= until)

    rows = q.order_by(AuditLog.created_at.desc()).limit(limit).all()

    # Build response dicts with resolved email addresses
    result = []
    for row in rows:
        admin_user  = db.query(User).filter(User.id == row.admin_id).first()  if row.admin_id       else None
        target_user = db.query(User).filter(User.id == row.target_user_id).first() if row.target_user_id else None
        result.append(AuditLogRow(
            id=row.id,
            admin_email=admin_user.email if admin_user else None,
            target_email=target_user.email if target_user else None,
            action=row.action,
            detail=row.detail,
            request_ip=row.request_ip,
            created_at=row.created_at,
        ))

    return AuditLogListResponse(logs=result)


# ---------------------------------------------------------------------------
# GET /admin/audit-logs/export  – download audit logs as Excel
# ---------------------------------------------------------------------------

_AUDIT_HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_AUDIT_HEADER_FILL  = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
_AUDIT_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_AUDIT_THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_AUDIT_EXPORT_HEADERS = ["ID", "Time", "User", "Action", "Request IP", "Details"]


@router.get("/audit-logs/export")
def export_audit_logs(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Export all audit logs as an Excel file.
    """
    rows = db.query(AuditLog).order_by(AuditLog.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    # Header row
    ws.append(_AUDIT_EXPORT_HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = _AUDIT_HEADER_FONT
        cell.fill = _AUDIT_HEADER_FILL
        cell.alignment = _AUDIT_HEADER_ALIGN
        cell.border = _AUDIT_THIN_BORDER

    # Data rows
    for row in rows:
        admin_user = db.query(User).filter(User.id == row.admin_id).first() if row.admin_id else None
        target_user = db.query(User).filter(User.id == row.target_user_id).first() if row.target_user_id else None
        user_email = (admin_user.email if admin_user else None) or (target_user.email if target_user else None) or ""

        ws.append([
            row.id,
            row.created_at.strftime("%Y-%m-%d %H:%M:%S") if row.created_at else "",
            user_email,
            row.action,
            row.request_ip or "",
            row.detail or "",
        ])
        # Apply border to every cell
        row_idx = ws.max_row
        for col_idx in range(1, len(_AUDIT_EXPORT_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _AUDIT_THIN_BORDER

    # Column widths
    _AUDIT_COL_MIN = [8, 20, 28, 18, 16, 50]
    for col_idx, min_w in enumerate(_AUDIT_COL_MIN, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = min_w

    # Stream
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="audit-logs.xlsx"'},
    )
